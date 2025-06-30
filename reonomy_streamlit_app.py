# --- Imports ---
import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import time
import math
import re
from io import BytesIO

# --- (1) Setup Selenium ---
def start_browser():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)  # Keeps browser open
    return webdriver.Chrome(options=options)

# --- (2) Scraper Loop ---
def scrape_all_pages_with_clicks(driver, num_pages):
    all_data = []
    headers = []

    for page in range(1, num_pages + 1):
        st.write(f"➡️ Scraping page {page} of {num_pages}")
        time.sleep(2)

        rows = driver.find_elements(By.TAG_NAME, "tr")
        st.write(f"   Found {len(rows)} rows")

        if not headers:
            for row in rows:
                ths = row.find_elements(By.TAG_NAME, "th")
                if ths:
                    headers = [th.text.strip() for th in ths]
                    break

            st.write("Extracted headers:", headers)

        for row in rows:
            tds = row.find_elements(By.TAG_NAME, "td")
            ths = row.find_elements(By.TAG_NAME, "th")
            if len(tds) >= 3:
                row_data = []
                row_number = tds[0].text.strip()
                row_data.append(row_number)

                address_links = ths[0].text.strip() if ths else ""
                row_data.append(address_links)

                for td in tds[1:]:
                    row_data.append(td.text.strip())

                all_data.append(row_data)

        # Click next page if not last
        if page < num_pages:
            try:
                next_button = driver.find_element(By.XPATH, f"//span[text()='{page + 1}']/parent::button")
                driver.execute_script("arguments[0].click();", next_button)
            except Exception as e:
                st.warning(f"⚠️ Couldn't click to page {page + 1}: {e}")
                break

    if not all_data:
        st.error("⚠️ No data found.")
        return pd.DataFrame()

    df = pd.DataFrame(all_data, columns=headers if headers else None)
    return df


# --- (3) Get Total Pages ---
def get_total_pages(driver):
    try:
        element = driver.find_element(By.XPATH, "//a[@data-testid='company-properties-tab']//span[contains(text(), 'Properties')]")
        text = element.text  # e.g. "Properties (4,457)"
        match = re.search(r'\(([\d,]+)\)', text)
        if match:
            property_count = int(match.group(1).replace(",", ""))
            st.success(f"✅ Number of Properties: {property_count}")
            return math.ceil(property_count / 50)
    except Exception as e:
        st.error(f"❌ Couldn't extract total property count. Error: {e}")
    raise e

   
# --- (4) Clean Columns ---
def clean_dataframe(df):
    address_col = next((col for col in df.columns if "Address" in col and "Reported" not in col), None)
    if address_col:
        df["State"] = df[address_col].apply(lambda x: x.split(",")[2].strip().split()[0] if isinstance(x, str) and len(x.split(",")) >= 3 else "")

    if "Lot Size (SF)" in df.columns:
        df["Lot Size (SF)"] = pd.to_numeric(
            df["Lot Size (SF)"]
            .str.replace(" SF", "", regex=False)
            .str.replace(",", "", regex=False),
            errors="coerce"
        )
    return df


# --- (5) Excel ---

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO

def format_and_export_excel(df, company_name="Company", filename="reonomy_property_list.xlsx"):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from copy import copy
    from io import BytesIO

    # Save raw DataFrame to BytesIO with initial spacing
    output = BytesIO()
    df.to_excel(output, index=False, startrow=1, startcol=1)  # put table starting on row 10
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    ws.title = "Property List"
    ws.sheet_view.showGridLines = False


    # --- Insert Spacer Rows (to be safe)
    ws.insert_rows(1)
    ws.insert_rows(2)
    ws.insert_rows(3)
    ws.insert_rows(4)

    # --- Write Title Section (now shifted down + right)
    ws["B2"] = f"{company_name} Property List"
    ws["B2"].font = Font(size=16, bold=True)

    ws["B3"] = "Project: Reonomy Web Scraper"
    ws["B3"].font = Font(size=13)

    ws["B4"] = "(All #'s in Actuals)"
    ws["B4"].font = Font(size=11, italic=True)

    # --- Move all data rows (starting from row 10) up to row 6
    # Set expected column headers
    df.columns = [
        "#",
        "Address",
        "Lot Size (SF)",
        "Contact Info Available",
        "Reported Owner",
        "Reported Owner Address",
        "State"
    ]

    # Export starting at row 6, column B
    output = BytesIO()
    df.to_excel(output, index=False, startrow=5, startcol=1)
    output.seek(0)

    # --- Style header row (row 6)
    for cell in ws[6]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Autofit columns
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Clear what was originally row 10 (now empty)
    for cell in ws[10]:
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.alignment = Alignment()
        cell.border = Border()

    # --- Continue with Ideal Addresses copy + filtering (same as before)
    if "Ideal Addresses" in wb.sheetnames:
        del wb["Ideal Addresses"]
    ws_target = wb.copy_worksheet(ws)
    ws_target.title = "Ideal Addresses"
    ws_target.sheet_view.showGridLines = False

    for row in ws_target.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("Project:"):
                cell.value = "Project: Properties ≥ 100,000 SF"
                break

    header_row_idx, lot_size_col_idx = None, None
    for row in ws_target.iter_rows(min_row=1, max_row=15):
        values = [cell.value for cell in row]
        if "Lot Size (SF)" in values:
            header_row_idx = row[0].row
            lot_size_col_idx = values.index("Lot Size (SF)") + 1
            break

    if header_row_idx and lot_size_col_idx:
        rows_to_keep = []
        for row in ws_target.iter_rows(min_row=header_row_idx + 1, max_row=ws_target.max_row):
            val = row[lot_size_col_idx - 1].value
            try:
                if val and float(val) >= 100000:
                    rows_to_keep.append([cell.value for cell in row])
            except:
                continue

        ws_target.delete_rows(header_row_idx + 1, ws_target.max_row - header_row_idx)
        for i, row_data in enumerate(rows_to_keep, start=header_row_idx + 1):
            for j, value in enumerate(row_data, start=1):
                ws_target.cell(row=i, column=j, value=value)

    # --- Add double border and spacing rows to both sheets
    double_bottom = Border(bottom=Side(border_style="double", color="000000"))

    for sheetname in ["Property List", "Ideal Addresses"]:
        wsx = wb[sheetname]
        last_row = 0
        for row in wsx.iter_rows(min_row=1, max_row=wsx.max_row):
            if any(cell.value not in [None, ""] for cell in row):
                last_row = row[0].row

        spacer1, spacer2 = last_row + 1, last_row + 2
        wsx.row_dimensions[spacer1].height = 4
        wsx.row_dimensions[spacer2].height = 4
        for col in range(1, 8):
            cell = wsx.cell(row=spacer1, column=col)
            cell.border = double_bottom

    # --- Clear column A in both sheets ---
    for sheetname in ["Property List", "Ideal Addresses"]:
        wsx = wb[sheetname]

        # Remove values and formatting in column A
        for cell in wsx["A"]:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.border = Border()

        # Collapse column A visually
        wsx.column_dimensions["A"].width = 2


    # --- Return file to Streamlit
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output


# --- (5) Streamlit UI ---
st.set_page_config(page_title="Reonomy Scraper", layout="wide")
st.title("🔍 Reonomy Property Scraper")

if 'driver' not in st.session_state:
    st.session_state.driver = None

# Step 1: Launch Chrome Browser for Manual Login
if st.button("1️⃣ Start Chrome Browser"):
    st.session_state.driver = start_browser()
    st.session_state.driver.get("https://app.reonomy.com/login")
    st.info("🧠 Please log in manually in the browser. Once logged in and on the company’s property page, paste the URL below.")

# Step 2: User Inputs Company URL
if st.session_state.driver:
    company_url = st.text_input("2️⃣ Paste Reonomy Company Properties URL:")

    # Step 3: Scrape Button
    if st.button("3️⃣ Scrape Properties"):
        if company_url:
            driver = st.session_state.driver
            driver.get(company_url)
            time.sleep(5)

            try:
                company_name = driver.find_element(By.CSS_SELECTOR, "h6.MuiTypography-subtitle2").text.strip()
            except:
                st.error("❌ Could not extract company name.")
                company_name = "company"

            total_pages = get_total_pages(driver)

            if total_pages > 0:
                with st.spinner("🔄 Scraping in progress..."):
                    df = scrape_all_pages_with_clicks(driver, total_pages)
                    df = clean_dataframe(df)

                st.success(f"✅ Scraped {len(df)} properties for **{company_name}**")
                st.dataframe(df)

                # CSV Download
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📥 Download CSV",
                    data=csv,
                )
                # Excel Download with header formatting
                excel_bytes = format_and_export_excel(df, filename=f"{company_name}_properties.xlsx")
                st.download_button(
                    label="📥 Download Formatted Excel",
                    data=excel_bytes,
                    file_name=f"{company_name}_properties.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        else:
            st.warning("⚠️ Please enter a valid company properties URL.")


#streamlit run reonomy_streamlit_app.py